import numpy as np
import openpyxl as xl
import pandas as pd
import pytest
from test_utils import *

from sql2excel.config import Config
from sql2excel.chart import Chart
from sql2excel.excel_helper import ExcelHelper
from sql2excel.sqlexec import SQLExecutor


@pytest.fixture(scope="session", autouse=True)
def workbook():
    wb = xl.Workbook()
    yield wb
    wb.close()


@pytest.fixture(scope="session", autouse=True)
def worksheet(workbook):
    ws = workbook.active
    return ws


@pytest.fixture(scope="session")
def config():
    return Config()


@pytest.fixture(scope="session", autouse=True)
def excel_helper(config):
    excel_helper = ExcelHelper(config=config)
    return excel_helper


@pytest.fixture(scope="function", autouse=True)
def sql_executor():
    return SQLExecutor(conn="dummy")


@pytest.fixture(scope="session", autouse=True)
def chart(config, excel_helper):
    return Chart(config=config, excel_helper=excel_helper)


@pytest.fixture(scope="session", autouse=True)
def df():
    n_rows = 10
    data = {
        "pubyear": np.arange(2000, 2000 + n_rows),
        "pubcount": np.random.randint(100, 1000, size=n_rows),
        "perc": np.random.rand(n_rows),
    }

    return pd.DataFrame(data)


@pytest.fixture(scope="function")
def random_dataframes():
    def _generate_dataframes(n):
        dataframes = []
        for _ in range(n):
            rows = np.random.randint(3, 10)
            cols = np.random.randint(2, 5)
            df = pd.DataFrame(
                np.random.randint(0, 100, size=(rows, cols)),
                columns=[f"Col{i+1}" for i in range(cols)],
            )
            dataframes.append(df)
        return dataframes

    return _generate_dataframes


@pytest.fixture(scope="session", autouse=True)
def long_wide_df():
    n_rows = 10
    n_cols = 6
    data = np.random.rand(n_rows, n_cols - 1)
    columns = [f"Column_{i+1}" for i in range(n_cols - 1)]
    df = pd.DataFrame(data, columns=columns)
    df.insert(
        0,
        column="pubyear",
        value=np.arange(2000, 2000 + n_rows),
    )
    return df


@pytest.fixture(scope="function")
def create_worksheet(workbook):
    return workbook.create_sheet()


@pytest.fixture(scope="function")
def clear_worksheet(worksheet):
    while worksheet.max_row > 1:
        worksheet.delete_rows(2)

    worksheet.delete_rows(1)
    return worksheet


@pytest.fixture(scope="function")
def add_content_to_sheet(worksheet):
    n_rows = np.random.randint(1, 11)
    for r in range(n_rows):
        worksheet.append([f"Row #{r}"])

    return worksheet
