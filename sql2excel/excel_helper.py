"""
A module for manipulating Excel worksheets and charts using openpyxl.

Provides ExcelHelper which contains helper methods for formatting cells, working
with worksheet and customizing chart appearance (fonts, colors, axis limits, etc.)
"""

import warnings

import openpyxl as xl
import openpyxl.drawing.colors as colors

# from openpyxl.chart.series import Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import CharacterProperties
from openpyxl.drawing.text import Font as DrawingFont
from openpyxl.drawing.text import Paragraph, ParagraphProperties, RichTextProperties
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from sql2excel.config import Config


class ExcelHelper:
    def __init__(self, config=Config()) -> None:
        self.config = config

    def is_sheet_empty(self, ws: Worksheet):
        # NOTE This will return True if the sheet is iterated upon even though it contains "empty" cells
        return len(list(ws.rows)) == 0

        # Check if sheet is empty even if it was iterated upon
        # for row in ws.iter_rows():
        #     if any(cell.value is not None for cell in row):
        #         return False
        # return True

    def get_row_start(self, ws: Worksheet):
        if self.is_sheet_empty(ws):
            rowstart = 1
        else:
            # rowstart = len(list(ws.rows)) + 1
            rowstart = ws.max_row + self.config.SEPARATOR + 1
            for _ in range(self.config.SEPARATOR):
                ws.append([None])
        return rowstart

    def get_column_letter(self, col: str | int | None) -> str:
        if col is None:
            return "A"
        elif isinstance(col, str):
            col_index = xl.utils.cell.column_index_from_string(col)
            # max number of columns = 16384
            if col_index > 16384:
                # Silently reset the column if invalid
                col = "A"
                warnings.warn(
                    "Invalid column. Using column 'A' instead", category=UserWarning
                )
            return col
        elif isinstance(col, int) and col > 0 and col <= 16384:
            return xl.utils.cell.get_column_letter(col)
        else:
            warnings.warn(
                "Invalid column. Using column 'A' instead", category=UserWarning
            )
            return "A"

    def get_starting_position(self, ws, **kwargs):
        # Extract and validate row_start
        row_start = kwargs.get("row_start")
        if row_start is not None:
            if not isinstance(row_start, int) or row_start <= 0:
                row_start = None
                warnings.warn(
                    "Row start must be an integer >= 1.Choosing row number"
                    "dynamically depending on the data already in the sheet"
                )

        row_start = row_start or self.get_row_start(ws)

        # column start
        column_letter = self.get_column_letter(col=kwargs.get("column_start"))
        row_start, column_start = xl.utils.cell.coordinate_to_tuple(
            column_letter + str(row_start)
        )

        return row_start, column_start

    def insert_rows_for_chart_height(self, height, ws, df=None, scale=None):
        """Add empty rows to account for chart/image height"""
        scale = scale or self.config.CHART_HEIGHT_SCALE
        # Number of lines consumed by the chart
        n_rows = int(scale * height) + 1

        # Number of rows already occupied by dataframe
        n_rows_df = 0 if df is None else df.shape[0]

        # Number of rows to append: do not append if negative
        n_rows = n_rows - n_rows_df

        for _ in range(n_rows):
            ws.append([None])

    def set_line_graphical_properties(
        self,
        series: xl.chart.series.Series,
        width: float = None,
        style: str = None,
        color: str = None,
        smooth: bool = None,
        nofill: bool = False,
    ) -> None:
        """
        Set graphic properties for a chart series.
        """

        if nofill:
            series.graphicalProperties.line.noFill = nofill
            return

        if style:
            series.graphicalProperties.line.dashStyle = style

        if width:
            # 1 point = 12700 EMU
            series.graphicalProperties.line.width = 12700 * width

        if color:
            color = colors.ColorChoice(srgbClr=color)
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color

        if smooth is not None:
            series.smooth = smooth

    def set_marker_graphical_properties(
        self,
        series: xl.chart.series.Series,
        symbol: str = None,
        size: int = None,
        color: str = None,
    ) -> None:
        """
        Set markers' graphic properties for a chart series.
        """

        if symbol:
            series.marker.symbol = symbol
            series.marker.size = size or self.config.MARKER_SIZE

        if color:
            color = colors.ColorChoice(srgbClr=color)
            series.marker.graphicalProperties.solidFill = color
            series.marker.graphicalProperties.line.solidFill = color

    def fill(
        self,
        series: xl.chart.series.Series,
        color: str = None,
        border_line_color=None,
    ) -> None:
        """
        Set graphic properties for a chart series.

        Parameters:
        -----------
        series : ChartSeries
            The chart series to customize.
        color : str, optional
            The color (srgbClr) of the line and marker. Defaults to the value in the configuration.

        Returns:
        --------
        None
        """

        color = colors.ColorChoice(srgbClr=color)
        if color:
            series.graphicalProperties.solidFill = color

            border_line_color = (
                colors.ColorChoice(srgbClr=border_line_color)
                if border_line_color
                else color
            )
            series.graphicalProperties.line.solidFill = border_line_color
            # TODO change method name and allow for setting line properties

    def fill_data_point(self, series, series_length):
        # NOTE Accessing series length requires referencing the worksheet and finding the cell range
        # It is easier to let the user provides this
        data_points = []
        for idx in range(series_length):
            pt = xl.chart.marker.DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = xl.drawing.colors.ColorChoice(
                # Recycle colors if needed
                srgbClr=self.config.PRIMARY_COLORS[
                    idx % len(self.config.PRIMARY_COLORS)
                ]
            )
            data_points.append(pt)

            series.data_points = data_points

    def reference_column_exists(self, df):
        return df.iloc[:, -1].nunique() == 1

    def set_section_heading_font(
        self,
        cell,
        sh_font_name=None,
        sh_bold=None,
        sh_font_size=None,
        sh_font_color=None,
        **kwargs,
    ):
        sh_font_name = sh_font_name or self.config.SECTION_HEADING_FONT_NAME
        sh_bold = sh_bold or self.config.SECTION_HEADING_BOLD or sh_bold
        sh_font_size = sh_font_size or self.config.SECTION_HEADING_FONT_SIZE
        sh_font_color = sh_font_color or self.config.SECTION_HEADING_FONT_COLOR

        cell.font = Font(
            name=sh_font_name,
            bold=sh_bold,
            size=sh_font_size,
            color=sh_font_color,
        )

    def set_df_title_font(
        self,
        cell,
        df_font_name=None,
        df_bold=None,
        df_font_size=None,
        df_font_color=None,
        **kwargs,
    ):
        # Use provided arguments or fall back to configuration defaults
        df_font_name = df_font_name or self.config.DF_TITLE_FONT_NAME
        df_bold = df_bold or self.config.DF_TITLE_BOLD
        df_font_size = df_font_size or self.config.DF_TITLE_FONT_SIZE
        df_font_color = df_font_color or self.config.DF_TITLE_FONT_COLOR

        # Apply the font style to the cell
        cell.font = Font(
            name=df_font_name,
            bold=df_bold,
            size=df_font_size,
            color=df_font_color,
        )

    def set_chart_title_font(self, chart_, **kwargs):
        # Do not remove underscore from `chart_` as kwargs might contain a key 'chart'.
        # A query in SQL script can be annotated with 'chart' to export the result
        if chart_.title is None:
            return

        font_name = DrawingFont(typeface=self.config.CHART_TITLE_FONT_NAME)
        if kwargs.get("title_font_name"):
            font_name = DrawingFont(typeface=kwargs.get("title_font_name"))
        font_size = kwargs.get("title_font_size") or self.config.CHART_TITLE_FONT_SIZE
        color = kwargs.get("title_font_color") or self.config.CHART_TITLE_FONT_COLOR
        bold = kwargs.get("title_font_bold") or self.config.CHART_TITLE_FONT_BOLD

        color = (
            colors.ColorChoice(prstClr=color)
            if color in self.config.VALID_COLORS
            else colors.ColorChoice(srgbClr=color)
        )

        cp = xl.drawing.text.CharacterProperties(
            latin=font_name,
            sz=font_size,
            solidFill=color,
            b=bold,
        )

        try:
            if chart_.title:
                chart_.title.tx.rich.p[0].r[0].rPr = cp
        except (IndexError, AttributeError):
            warnings.warn(
                "Unable to set chart title", category=UserWarning, stacklevel=1
            )

    def set_chart_axis_label_font(self, chart_, axis, **kwargs):
        # Do not remove underscore from `chart_` as kwargs might contain a key 'chart'.
        # A query in SQL script can be annotated with 'chart' to export the result

        axis_font_name = kwargs.get("axis_font_name") or self.config.AXIS_FONT_NAME
        axis_font_size = kwargs.get("axis_font_size") or self.config.AXIS_FONT_SIZE
        axis_font_color = kwargs.get("axis_font_color") or self.config.AXIS_FONT_COLOR
        axis_font_bold = kwargs.get("axis_font_bold") or self.config.AXIS_FONT_BOLD

        font = DrawingFont(typeface=axis_font_name)
        axis_font_color = (
            colors.ColorChoice(prstClr=axis_font_color)
            if axis_font_color in self.config.VALID_COLORS
            else colors.ColorChoice(srgbClr=axis_font_color)
        )
        cp = xl.drawing.text.CharacterProperties(
            latin=font,
            sz=axis_font_size,
            solidFill=axis_font_color,
            b=axis_font_bold,
        )

        try:
            if axis.lower() == "x" and chart_.x_axis.title is not None:
                chart_.x_axis.title.tx.rich.p[0].r[0].rPr = cp
            elif axis.lower() == "y" and chart_.y_axis.title is not None:
                chart_.y_axis.title.tx.rich.p[0].r[0].rPr = cp
            else:
                pass
        except AttributeError:
            print(chart_.__class__)
            warnings.warn("Unable to style axis label", category=UserWarning)

    def rotate_xticks(self, chart, rotation):
        """Rotate the xtick label of the x-axis"""

        rotation *= 60000

        try:
            chart.x_axis.txPr = RichText(
                bodyPr=RichTextProperties(
                    anchor="ctr",
                    anchorCtr="1",
                    rot=rotation,
                    spcFirstLastPara="1",
                    vertOverflow="ellipsis",
                    wrap="square",
                ),
                p=[
                    Paragraph(
                        pPr=ParagraphProperties(defRPr=CharacterProperties()),
                        endParaRPr=CharacterProperties(),
                    )
                ],
            )
        except:
            warnings.warn(
                "Unable to set rotate xticks", category=UserWarning, stacklevel=1
            )

    def set_axis_limit(self, axis, limit):
        mini, maxi = limit
        if mini >= maxi:
            warnings.warn(
                f"Invalid axis limits: {limit}. Axis limits will be ignored",
                category=UserWarning,
                stacklevel=1,
            )
        axis.scaling.min = mini
        axis.scaling.max = maxi


if __name__ == "__main__":
    pass
