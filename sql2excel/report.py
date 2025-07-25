"""
Module for generating Excel reports from SQL queries with charting support.

This module provides the Report class, which executes SQL queries, processes results,
and writes them to Excel files, optionally including charts. It integrates with pandas,
openpyxl, and custom chart classes.
"""

# TODO logging problematic queries - Do not crash

from typing import Sequence

import openpyxl as xl
import pandas as pd
from sql2excel.config import Config
from sql2excel.chart import (
    AreaChart,
    BarChart,
    BarLineChart,
    BubbleChart,
    Chart,
    LineChart,
    PieChart,
    RadarChart,
    ScatterChart,
    StackedBarChart,
)
from sql2excel.excel_helper import ExcelHelper
from sql2excel.sqlexec import QueryConfig, SQLExecutor

CHART_MAP = {
    "chart": Chart,
    "area": AreaChart,
    "line": LineChart,
    "bar": BarChart,
    "barline": BarLineChart,
    "pie": PieChart,
    "radar": RadarChart,
    "bubble": BubbleChart,
    "scatter": ScatterChart,
    "stackedbar": StackedBarChart,
}


class Report:

    def __init__(
        self,
        conn=None,
        session=None,
        engine=None,
        connection_string=None,
        config=None,
        silent=False,
    ) -> None:
        self.config = config or Config()
        self.excel_helper = ExcelHelper(config=self.config)
        self.executor = SQLExecutor(
            conn=conn,
            session=session,
            engine=engine,
            connection_string=connection_string,
            silent=silent,
        )

    def generate(self, query_config: QueryConfig | Sequence[QueryConfig], **kwarg):

        fname = kwarg.get("file_name", "result.xlsx")
        wb = xl.Workbook()
        ws = wb.active
        sheetname = kwarg.get("sheetname")
        if sheetname:
            ws.title = sheetname

        if isinstance(query_config, QueryConfig):
            queries_config = [query_config]
        else:
            queries_config = query_config

        # results = self.executor.executeall(queries_config)
        query_results = []
        for qc in queries_config:
            if qc.from_sql_script and "chart" not in qc.xl_params.keys():
                continue

            df = self.executor.execute(qc)

            if df is None:
                continue

            index = qc.xl_params.get("index")
            if index:
                columns = qc.xl_params.get("columns")
                values = qc.xl_params.get("values")
                df = pd.pivot(
                    df, index=index, columns=columns, values=values
                ).reset_index(drop=False)

            query_results.append((qc, df))

        for idx, (qc, df) in enumerate(query_results):

            sheetname = qc.xl_params.get("sheetname") or qc.xl_params.get("sheet name")
            sheetname = sheetname.strip() if sheetname else None
            if sheetname:
                if idx == 0:
                    del wb[ws.title]

                if sheetname not in wb.sheetnames:
                    ws = wb.create_sheet(sheetname)
                else:
                    ws = wb[sheetname]

            chart = CHART_MAP.get(qc.xl_params.get("chart"))
            if chart:
                chart = chart(config=self.config, excel_helper=self.excel_helper)

                if chart.__class__.__name__ == "Chart":
                    chart.write_dataframe(df, ws, **qc.xl_params)
                else:
                    chart.plot(df, ws, **qc.xl_params)

        self.executor.close()

        wb.save(fname)

    def close(self):
        self.executor.close()


if __name__ == "__main__":
    pass
