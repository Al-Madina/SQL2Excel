"""
This module supports writing Pandas DataFrame to Excel including chart creation.

Classes
-------
Chart
    Base class for all chart types, providing data writing, chart generation, and customization.
LineBasedChart
    Abstract class for line and radar charts.
LineChart
    A concrete class for line chart.
RadarChart
    A concrete class for radar (spider) chart implementation.
BarChart
    A concrete class for bar chart implementation with color and axis customization.
StackedBarChart
    A concrete class for stacked and percent-stacked bar chart implementation.
PieChart
    A concrete class for pie chart with data label customization.
AreaChart
    A concrete class for area chart implementation.
ScatterChart
    A concrete class for scatter plot implementation with flexible data selection.
BubbleChart
    A concrete class for bubble chart implementation using x, y, and size columns.
ImageChart
    A concrete class for to for embedding images or matplotlib figures into Excel.
TwoAxesChart
    Base class for chart with two y-axes (primary and secondary).
BarLineChart
    A concrete class for combined bar and line chart with dual y-axes.

Notes
-----
- Subclasses should override `plot` and `add_image` as needed for specific chart logic.
- All chart classes support basic customization via keyword arguments and/or configuration objects.
"""

import io
import warnings
from typing import Sequence

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

from sql2excel.config import Config
from sql2excel.excel_helper import ExcelHelper


class Chart:
    """
    Base class for all charts in SQL2Excel.

    This class provides the methods and attributes for creating, configuring,
    and rendering charts in Excel workbooks using pandas DataFrames and openpyxl.

    Parameters:
    -----------
    config : Config
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper
            Helper object for Excel-specific operations.

    Notes
    -----
    - If Excel-related code fails, a user warning is issued and a default value
    is used instead of raising an exception. This design choice is to prevent
    interrupting SQL queries on the failure of Excel-related code.
    - Subclasses should override `plot` and `add_image` to implement specific
    charting and image insertion logic.
    """

    def __init__(self, config=Config(), excel_helper=None):
        self.config = config
        self.excel_helper = excel_helper or ExcelHelper(self.config)

        # Chart: this must be initialized in `plot` method which should be overriden
        self.chart: xl.chart.Chart = None

        # Data range
        self.min_row = -1
        self.min_col = -1
        self.max_row = -1
        self.max_col = -1

        # Indexes of columns utilized for chart generation
        # This is particularly useful when the DataFrame contains multiple
        # columns, but only a subset is used for the chart
        self.data_columns = None

        # Reference column
        # NOTE 'reference_column' is the column contain a single unique value (baseline)
        self.ref_series_idx = None

    def write_dataframe(
        self, df: pd.DataFrame, ws: xl.worksheet.worksheet.Worksheet, **kwargs
    ):
        """Writes a pandas DataFrame to an Excel worksheet.

        Parameters
        ----------
        df : pd.DataFrame
            The DataFrame to write to the worksheet.
        ws : openpyxl.worksheet.worksheet.Worksheet
            The openpyxl worksheet where the DataFrame will be written.
        **kwargs : dict, optional
            Additional keyword arguments:
                section_heading : str, optional
                    A heading to be written above the DataFrame in the worksheet.
                headings : iterable of str, optional
                    Custom column headings to use instead of DataFrame's columns.
                Any other keyword arguments required by `self.excel_helper.get_starting_position`.

        Notes
        -----
        - The first column is always treated as the x-axis (categories).
        - If `headings` is provided, DataFrame's column names are ignored.
        """

        # Chart position
        row_start, column_start = self.excel_helper.get_starting_position(ws, **kwargs)

        # A heading describing the data
        section_heading = kwargs.get("section_heading")
        if section_heading:
            ws.cell(row=row_start, column=column_start, value=section_heading)

            self.excel_helper.set_section_heading_font(
                ws.cell(row=row_start, column=column_start)
            )

            # Update
            row_start += 1

        r, c = df.shape
        # `row_min` includes the headings
        self.min_row = row_start
        # First column is always the x-axis (categories) - Not a part of the data range
        self.min_col = column_start + 1
        self.max_row = row_start + r
        self.max_col = column_start + c - 1

        # Column headings should be an iterable of str
        headings = kwargs.get("headings")
        if headings:
            for idx, heading in enumerate(headings):
                ws.cell(row=row_start, column=column_start + idx, value=heading)

            row_start += 1

            # Ignore column names in df
            rows = dataframe_to_rows(df, index=False, header=False)
        else:
            rows = dataframe_to_rows(df, index=False, header=True)

        for row_idx, row in enumerate(rows):
            for col_idx in range(len(row)):
                ws.cell(
                    row=row_start + row_idx,
                    column=column_start + col_idx,
                    value=row[col_idx],
                )

    def write_dataframes_side_by_side(
        self,
        objs: Sequence[pd.DataFrame],
        ws: xl.worksheet.worksheet.Worksheet,
        **kwargs,
    ):
        """
        Write multiple pandas DataFrames side by side into an Excel worksheet.

        Parameters
        ----------
        objs : Sequence[pd.DataFrame]
            Sequence of pandas DataFrames to be written to the worksheet.
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet where the DataFrames will be written.
        **kwargs : dict, optional
            Additional keyword arguments:
                section_heading : str, optional
                    Title to add above the DataFrames section.
                df_headings : list of str, optional
                    Custom headings for each DataFrame. The length should match the number of DataFrames in `objs`.
                row_start : int, optional
                    The starting row for writing DataFrames. Defaults to the value returned by `get_starting_position`.
                column_start : int, optional
                    The starting column for writing DataFrames. Defaults to 1 (the first column).
                headings : list of list of str, optional
                    Custom column headings for each DataFrame. Each element should be a list of headings for the corresponding DataFrame.
                Other keyword arguments are passed to the underlying DataFrame writing method.

        See Also
        --------
        write_dataframe : Method used to write individual DataFrames.

        Notes
        -----
        - DataFrames are written side by side, separated by a configurable number of columns.
        - The method delegates the actual writing of each DataFrame to `write_dataframe`.
        """

        row_start, column_start = self.excel_helper.get_starting_position(ws, **kwargs)

        # A heading describing the data
        section_heading = kwargs.get("section_heading")
        if section_heading:
            ws.cell(row=row_start, column=column_start, value=section_heading)

            self.excel_helper.set_section_heading_font(
                ws.cell(row=row_start, column=column_start)
            )

            # Delete because `write_dataframe` will be delegated for writing individual df
            del kwargs["section_heading"]

            # Update
            row_start += 1

        df_headings = kwargs.get("df_headings")
        headings_seq = kwargs.get("headings")
        if headings_seq:
            del kwargs["headings"]

        n_columns = 0
        for idx, df in enumerate(objs):
            current_column = column_start + n_columns
            current_row = row_start
            if df_headings:
                try:
                    ws.cell(
                        row=row_start, column=current_column, value=df_headings[idx]
                    )

                    self.excel_helper.set_df_title_font(
                        ws.cell(row=row_start, column=current_column), **kwargs
                    )
                except IndexError:
                    warnings.warn(
                        "The number of DataFrame headings provided does not match the number of DataFrames",
                        category=UserWarning,
                        stacklevel=1,
                    )

                current_row += 1

            # Write the current dataframe at the specified position (current_row, current_column)
            headings = headings_seq[idx] if headings_seq else None
            self.write_dataframe(
                df,
                ws,
                row_start=current_row,
                column_start=current_column,
                headings=headings,
                **kwargs,
            )

            n_columns += len(df.columns) + self.config.DATA_DATA_SEPARATOR

    # To be overridden in subclasses
    def add_image(self, image_path: str, ws, df=None, **kwargs):
        """
        Adds an image to the given worksheet.

        Parameters
        ----------
        image_path : str
            The file path to the image to be inserted.
        ws : object
            The worksheet object where the image will be added.
        df : pandas.DataFrame, optional
            DataFrame that may be written alongside the image (default is None).
        **kwargs
            Additional keyword arguments for image customization, such as position,
            size, or formatting.

        Raises
        ------
        NotImplementedError
            If the method is not implemented in the subclass.

        Notes
        -----
        This method should be implemented by subclasses.
        """

        raise NotImplementedError("This method has not been implemented yet")

    def _add_data(self, df, ws, from_rows=False, set_categories=True, **kwargs):
        """
        Add data from a DataFrame to an Excel chart.

        Parameters
        ----------
        df : pandas.DataFrame
            The DataFrame containing the data to be added to the chart.
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet object where the data is located.
        from_rows : bool, default False
            If True, data is added from rows (wide format). Otherwise, data is
            added from columns (long format).
        set_categories : bool, default True
            If True, sets the chart categories using the first column of the data.
        **kwargs : dict, optional
            Additional keyword arguments:
                - data_column_start : int, optional
                    The starting column index for the data to be added.
                - data_column_end : int, optional
                    The ending column index for the data to be added.
                - data_columns : list of int, optional
                    Specific column indices to be added as data series.
                - use_ref_line : bool, optional
                    Whether to use a reference line for series with a single unique value.
                - column_start : int, optional
                    Offset for the starting column if data position is determined manually.

        Raises
        ------
        NotImplementedError
            If attempting to specify data range for data added from rows (wide format).
        UserWarning
            If data added from rows contains only one unique value in the last column.

        Notes
        -----
        - Categories are set using the first column by default.
        - When `from_rows` is True, only the default data range is supported.
        """
        # Add data to the chart
        if from_rows:
            if kwargs.get("data_column_start") or kwargs.get("data_columns"):
                raise NotImplementedError(
                    "This feature has not been implemented for data added from rows. To specify data range your data must be add in column format (long format)"
                )
            if df.iloc[:, -1].nunique() == 1:
                warnings.warn(
                    "This feature has not been implemetned for data added from rows.  To specify data range your data must be add in column format (long format)",
                    category=UserWarning,
                )
            return self._add_data_from_rows(ws, **kwargs)

        data_column_start = kwargs.get("data_column_start")
        data_column_end = kwargs.get("data_column_end")
        # If column start is given but not column end, set column end to column start
        data_column_end = data_column_end or data_column_start
        data_columns = kwargs.get("data_columns")
        use_ref_line = kwargs.get("use_ref_line", self.config.USE_REF_LINE)

        # Offset the data columns if the data position in the sheet is determind manually
        offset = (
            0 if kwargs.get("column_start") is None else kwargs.get("column_start") - 1
        )

        if data_columns is not None:
            data_columns = sorted(list(set(data_columns)))
            # Create each series separately and add it to the figure since columns are not adjacent
            for idx, col in enumerate(data_columns):
                values = xl.chart.Reference(
                    ws,
                    min_row=self.min_row,
                    max_row=self.max_row,
                    min_col=col + offset,
                    max_col=col + offset,
                )

                series = xl.chart.Series(
                    values=values,
                    title_from_data=True,
                )

                self.chart.series.append(series)

                # Find the index of reference column
                # NOTE Pandas index = openpyxl index - 1
                if df.iloc[:, col - 1].nunique() == 1 and use_ref_line:
                    self.ref_series_idx = idx

            if set_categories:
                categories = xl.chart.Reference(
                    ws,
                    min_row=self.min_row + 1,
                    max_row=self.max_row,
                    # `min_col` is the start of the data. The category is always the first column
                    min_col=self.min_col - 1,
                    max_col=self.min_col - 1,
                )

                self.chart.set_categories(categories)

            self.data_columns = list(data_columns)

            # if df.iloc[:, data_columns[-1] - 1].nunique() == 1:
            #     self.ref_col_idx = data_columns[-1]

            # Terminate to avoid adding data to the chart again
            return

        elif data_column_start is not None:
            data_reference = xl.chart.Reference(
                ws,
                min_row=self.min_row,
                max_row=self.max_row,
                min_col=data_column_start + offset,
                max_col=data_column_end + offset,
            )

            # Use pyxl indexing (starting from 1)
            self.data_columns = list(range(data_column_start, data_column_end + 1))

            for idx, col in enumerate(
                list(range(data_column_start, data_column_end + 1))
            ):
                if df.iloc[:, col - 1].nunique() == 1 and use_ref_line:
                    self.ref_series_idx = idx
        else:
            #
            data_reference = xl.chart.Reference(
                ws,
                min_row=self.min_row,
                max_row=self.max_row,
                min_col=self.min_col,
                max_col=self.max_col,
            )

            self.data_columns = list(range(self.min_col, self.max_col + 1))

            # Skip the first column as it is the category by default
            for idx in range(1, df.shape[1]):
                if df.iloc[:, idx].nunique() == 1 and use_ref_line:
                    # Series 0 is df column at index 1
                    self.ref_series_idx = idx - 1

        self.chart.add_data(data_reference, titles_from_data=True)

        if set_categories:
            # Add data and categories
            categories = xl.chart.Reference(
                ws,
                min_row=self.min_row + 1,
                max_row=self.max_row,
                # `min_col` is the start of the data. The category is always the first column
                min_col=self.min_col - 1,
                max_col=self.min_col - 1,
            )
            self.chart.set_categories(categories)

    # To be overridden in subclasses
    def plot(self, df, ws, **kwargs):
        """
        Plots data from a DataFrame onto a worksheet.

        Parameters
        ----------
        df : pandas.DataFrame
            The DataFrame containing the data to plot.
        ws : object
            The worksheet object where the plot will be rendered.
        **kwargs
            Additional keyword arguments for customizing the plot.

        Raises
        ------
        NotImplementedError
            If the method is not implemented.

        Notes
        -----
        This method should be implemented by subclasses.
        """

        raise NotImplementedError("This method have not been implemented yet")

    # To be called in by subclass's `plot` at the end to finalize the chart
    def _plot(self, df, ws, **kwargs):
        """
        Finalizes and adds a chart to an Excel worksheet.

        This method is intended to be called by subclass's `plot` method. It
        configures chart style, shape, titles, axis labels, axis limits, legend,
        chart dimensions, and adds the chart to the worksheet at the specified
        position. It also applies color fills to the chart series if required.

        Parameters
        ----------
        df : pandas.DataFrame
            The data to be plotted in the chart.
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet to which the chart will be added.
        **kwargs : dict, optional
            Additional keyword arguments to customize the chart:
                chart_style : int, optional
                    Style index for the chart.
                chart_shape : str, optional
                    Shape of the chart.
                title : str, optional
                    Title of the chart.
                xlabel : str, optional
                    Label for the x-axis.
                ylabel : str, optional
                    Label for the y-axis.
                xlim : tuple, optional
                    Limits for the x-axis.
                ylim : tuple, optional
                    Limits for the y-axis.
                rotation : int, optional
                    Rotation angle for x-axis tick labels.
                x_log_base : int, optional
                    Logarithmic base for the x-axis.
                y_log_base : int, optional
                    Logarithmic base for the y-axis.
                y_orientation : {'minMax', 'maxMin'}, optional
                    Orientation of the y-axis.
                yaxis_major_unit : int or float, optional
                    Major unit for the y-axis.
                width : int or float, optional
                    Width of the chart.
                height : int or float, optional
                    Height of the chart.
                show_legend : bool, optional
                    Whether to display the chart legend.
                legend_position : str, optional
                    Position of the legend.
                chart_position : {'bottom', 'right'}, optional
                    Position to insert the chart in the worksheet.
                suppress_chart_position_warning : bool, optional
                    Suppress warning if chart position is unknown.
                openpyxl_color : bool, optional
                    Use default Openpyxl colors for series.
                nofill : bool, optional
                    Do not fill series with colors.
                border_line_color : str, optional
                    Color for the border line of the series.
        """

        chart_style = kwargs.get("chart_style")
        if chart_style:
            self.chart.style = chart_style

        chart_shape = kwargs.get("chart_shape")
        if chart_shape:
            self.chart.shape = chart_shape

        # Title and labels
        title = kwargs.get("title")
        if title:
            self.chart.title = title
            self.excel_helper.set_chart_title_font(self.chart, **kwargs)

        # A chart can have no axes: Pie chart, etc
        # x-axis
        if hasattr(self.chart, "x_axis"):

            self.chart.x_axis.title = kwargs.get("xlabel")
            self.excel_helper.set_chart_axis_label_font(self.chart, axis="x", **kwargs)

            xlim = kwargs.get("xlim")
            if xlim:
                self.excel_helper.set_axis_limit(self.chart.x_axis, xlim)

            rotation = kwargs.get("rotation")
            # rotation can be 0
            rotation = (
                rotation if rotation is not None else self.config.XTICKST_ROTATION
            )
            self.excel_helper.rotate_xticks(self.chart, rotation)

            base = kwargs.get("x_log_base")
            if base:
                self.chart.x_axis.scaling.logBase = base

        # y axis
        if hasattr(self.chart, "y_axis"):
            self.chart.y_axis.title = kwargs.get("ylabel")
            self.excel_helper.set_chart_axis_label_font(self.chart, axis="y", **kwargs)

            ylim = kwargs.get("ylim")
            if ylim:
                self.excel_helper.set_axis_limit(self.chart.y_axis, ylim)

            # NOTE orientation does not work well on the y-axis
            # y-axis orientation
            # y_orientation = {'minMax', 'maxMin'}
            y_axis_orientation = kwargs.get("y_orientation")
            if y_axis_orientation:
                # orientation can be 'minMax' (normal y axis) or 'maxMin' (reverse y axis)
                self.chart.y_axis.scaling.orientation = y_axis_orientation

            # Axis Unit
            major_unit = kwargs.get("yaxis_major_unit")
            if major_unit:
                self.chart.y_axis.majorUnit = major_unit

            base = kwargs.get("y_log_base")
            if base:
                self.chart.y_axis.scaling.logBase = base

        # Chart dimension in this order: kwargs, config, default
        width, height = self.config.CHART_WIDTH, self.config.CHART_HEIGHT
        width = kwargs.get("width") or width or self.chart.width
        height = kwargs.get("height") or height or self.chart.height
        self.chart.width = width
        self.chart.height = height

        # Legend
        show_legend = kwargs.get("show_legend")  # True/False
        show_legend = self.config.SHOW_LEGEND if show_legend is None else show_legend
        if show_legend:
            legend_position = (
                kwargs.get("legend_position") or self.config.LEGEND_POSITION
            )
            self.chart.legend.position = legend_position

        else:
            self.chart.legend = None

        # Add chart to sheet
        chart_position = kwargs.get("chart_position") or self.config.CHART_POSITION

        if chart_position == "bottom":
            ref = xl.utils.cell.get_column_letter(self.min_col - 1) + str(
                self.max_row + self.config.DATA_CHART_SEPARATOR + 1
            )
            ws.add_chart(self.chart, ref)

            # Add empty rows after writing data and inserting the figure
            self.excel_helper.insert_rows_for_chart_height(height, ws, df=None)
        elif chart_position == "right":
            ref = xl.utils.cell.get_column_letter(
                self.max_col + self.config.DATA_CHART_SEPARATOR + 1
            ) + str(self.min_row)
            ws.add_chart(self.chart, ref)

            # Add empty rows after writing data and inserting the figure
            self.excel_helper.insert_rows_for_chart_height(height, ws, df)
        else:
            # NOTE do not try to resolve this by falling back to a default `chart_position`.
            # A chart might not be added on purpose. E.g. see `TwoAxesChart`.
            if not kwargs.get("suppress_chart_position_warning"):
                warnings.warn(
                    f"Unknown chart position: '{chart_position}'. Chart will not be added to the sheet",
                    category=UserWarning,
                    stacklevel=1,
                )
        # Fill in series with either default Openpyxl colors or SQL2Excel colors
        if (
            not self.config.OPENPYXL_COLORS
            and not kwargs.get("openpyxl_color")
            and not kwargs.get("nofill")
        ):
            for idx, series in enumerate(self.chart.series):
                # Recycle colors if required
                color = self.config.PRIMARY_COLORS[
                    idx % len(self.config.PRIMARY_COLORS)
                ]
                border_line_color = kwargs.get("border_line_color")
                self.excel_helper.fill(series, color, border_line_color)

    # TODO complete the implementation
    # TODO Handle reference
    # TODO Provide an option for specifying data range manually
    def _add_data_from_rows(self, ws, **kwargs):
        """
        Adds data from worksheet rows to the chart and sets the categories.

        Parameters
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet from which to reference data for the chart.
        **kwargs : dict
            Additional keyword arguments (currently unused).
        """

        categories = xl.chart.Reference(
            ws,
            min_row=self.min_row,
            max_row=self.min_row,
            min_col=self.min_col,
            max_col=self.max_col,
        )

        for row in range(self.min_row + 1, self.max_row + 1):
            values = xl.chart.Reference(
                ws,
                min_row=row,
                max_row=row,
                min_col=self.min_col - 1,
                max_col=self.max_col,
            )

            self.chart.add_data(values, titles_from_data=True, from_rows=True)

        self.chart.set_categories(categories)


class LineBasedChart(Chart):
    """
    A chart class for creating and customizing line-based charts.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.

    Notes
    -----
    Reference series are always formatted according to configuration..
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def _plot(self, df, ws, **kwargs):
        # Default 1.5
        line_width = kwargs.get("line_width", 1.5)
        line_style = kwargs.get("line_style")

        smooth = None
        # Radar chart has not attribute smooth
        if not isinstance(self.chart, xl.chart.RadarChart):
            smooth = kwargs.get("smooth", self.config.LINE_SMOOTH)

        nofill = kwargs.get("nofill")

        ref_found = 0
        for idx, series in enumerate(self.chart.series):
            if self.ref_series_idx is not None and idx == self.ref_series_idx:
                ref_found = 1
                continue

            color = (
                None  # Use default openpyxl color
                if self.config.OPENPYXL_COLORS or kwargs.get("openpyxl_color")
                # Recycle colors if needed
                else self.config.PRIMARY_COLORS[
                    (idx - ref_found) % len(self.config.PRIMARY_COLORS)
                ]
            )

            # Use line color if it is user-defined
            color = kwargs.get("line_color") or color
            self.excel_helper.set_line_graphical_properties(
                series, line_width, line_style, color, smooth, nofill
            )

            marker_symbol = kwargs.get("marker_symbol")
            marker_size = kwargs.get("marker_size")
            self.excel_helper.set_marker_graphical_properties(
                series, marker_symbol, marker_size, color
            )

        # Reference column is always set regardless of config.OPENPYXL_COLORS
        if self.ref_series_idx is not None:
            if isinstance(self.chart, xl.chart.RadarChart):
                self.excel_helper.set_line_graphical_properties(
                    self.chart.series[self.ref_series_idx],
                    self.config.RADAR_REF_WIDTH,
                    self.config.RADAR_REF_STYLE,
                    self.config.RADAR_REF_COLOR,
                    nofill=nofill,
                )
                # self.excel_helper.fill(
                #     self.chart.series[-1], self.config.RADAR_REF_COLOR
                # )
            elif isinstance(self.chart, xl.chart.LineChart):
                self.excel_helper.set_line_graphical_properties(
                    self.chart.series[self.ref_series_idx],
                    self.config.LINE_REF_WIDTH,
                    self.config.LINE_REF_STYLE,
                    self.config.LINE_REF_COLOR,
                    smooth=smooth,
                    nofill=nofill,
                )

        # Need to call the super method
        kwargs["nofill"] = True  # Preserve the current filling
        return super()._plot(df, ws, **kwargs)


class ImageChart(Chart):
    """
    A chart class for embedding images or matplotlib figures into Excel worksheets.
    This class provides methods to add images or matplotlib figures to an Excel worksheet,
    positioning them relative to dataframes or at specified locations, and configuring their
    size and placement.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def add_image(
        self,
        image_input: str | mpl.figure.Figure,
        ws,
        df=None,
        **kwargs,
    ):

        if isinstance(image_input, str):
            img = Image(image_input)
        else:
            img_bytes = io.BytesIO()

            if isinstance(image_input, mpl.figure.Figure):
                image_input.savefig(img_bytes, format="png", bbox_inches="tight")
                plt.close()
            else:
                warnings.warn(
                    "Unsupported image input type. No image will be added",
                    category=UserWarning,
                    stacklevel=1,
                )
                return

            img_bytes.seek(0)  # Rewind the BytesIO object to the beginning
            img = Image(img_bytes)

        if df is None:
            row_start, column_start = self.excel_helper.get_starting_position(
                ws, **kwargs
            )

            section_heading = kwargs.get("section_heading")
            if section_heading:
                ws.cell(row=row_start, column=column_start, value=section_heading)

                self.excel_helper.set_section_heading_font(
                    ws.cell(row=row_start, column=column_start)
                )

                # Update
                row_start += 1
        else:
            self.write_dataframe(df, ws, **kwargs)

        # Width and height in pixels
        height = kwargs.get("height") or self.config.IMAGE_HEIGHT or img.height
        img.height = height
        img.width = kwargs.get("width") or self.config.IMAGE_WIDTH or img.width

        chart_position = kwargs.get("chart_position") or self.config.CHART_POSITION

        if chart_position == "right":
            anchor_x = self.min_row if df is not None else row_start
            anchor_y = (
                self.max_col + self.config.DATA_CHART_SEPARATOR + 1
                if df is not None
                else column_start
            )
            ref = xl.utils.cell.get_column_letter(anchor_y) + str(anchor_x)
            ws.add_image(img, ref)
            # Add empty rows after writing data and inserting the figure
            self.excel_helper.insert_rows_for_chart_height(
                height, ws, df=df, scale=self.config.IMAGE_HEIGHT_UNIT
            )
        elif chart_position == "bottom":
            anchor_x = (
                self.max_row + self.config.DATA_CHART_SEPARATOR + 1
                if df is not None
                else row_start
            )
            anchor_y = self.min_col - 1 if df is not None else column_start
            ref = xl.utils.cell.get_column_letter(anchor_y) + str(anchor_x)
            ws.add_image(img, ref)
            # Add empty rows after writing data and inserting the figure
            self.excel_helper.insert_rows_for_chart_height(
                height, ws, df=None, scale=self.config.IMAGE_HEIGHT_UNIT
            )
        else:
            warnings.warn(
                "Unknown chart position. Chart will not be added to the sheet",
                category=UserWarning,
                stacklevel=1,
            )


# TODO Transparent filling when `self.chart.type='filled'`
class RadarChart(LineBasedChart):
    """
    RadarChart for plotting radar (spider) charts in Excel using pandas DataFrames.
    This class extends LineBasedChart to provide functionality for creating radar charts
    in Excel worksheets. It supports customization of chart type, axis scaling, and line width.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.RadarChart()

        # 'marker', 'filled', 'standard'
        self.chart.type = kwargs.get("chart_type", "standard")

        self._add_data(df, ws, **kwargs)

        radar_unit = kwargs.get("radar_unit")
        radar_unit_steps = (
            kwargs.get("radar_unit_steps") or self.config.RADAR_UNIT_STEPS
        )
        if radar_unit_steps:
            # TODO Fix the code below to select data columns when specified with data_column_start or data_columns
            radar_unit_steps = (
                np.max(df.iloc[:, 1:].values) - np.min(df.iloc[:, 1:].values)
            ) / radar_unit_steps

        unit = radar_unit or radar_unit_steps

        if unit:
            self.chart.y_axis.majorUnit = round(unit, 1) if unit < 1 else round(unit)

        mini = 0
        maxi = float(np.max(df.iloc[:, 1:].values))
        self.chart.y_axis.scaling = xl.chart.axis.Scaling(min=mini, max=maxi + 0.5)

        kwargs["line_width"] = kwargs.get("line_width") or self.config.RADAR_REF_WIDTH

        super()._plot(df, ws, **kwargs)


class LineChart(LineBasedChart):
    """
    A class for creating and plotting line charts in Excel worksheets.
    This class extends `LineBasedChart` and provides functionality to plot a pandas DataFrame as a line chart
    in an Excel worksheet using the provided configuration and Excel helper utilities.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.LineChart()

        self._add_data(df, ws, **kwargs)

        super()._plot(df, ws, **kwargs)


class BarChart(Chart):
    """
    A class for creating and customizing bar charts in Excel using data from a pandas DataFrame.
    This class extends the `Chart` base class and provides methods to plot bar charts with various customization options, such as varying bar colors, adjusting axis tick label positions, and handling negative values. The chart is rendered in an Excel worksheet using the provided Excel helper utilities.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.BarChart()

        self._add_data(df, ws, **kwargs)

        # Assign a different color to every individual bar
        vary_color = kwargs.get("vary_color") or self.config.BAR_CHART_VARYING_COLOR
        if vary_color and len(self.chart.series) == 1:
            kwargs["show_legend"] = False
            self.excel_helper.fill_data_point(self.chart.series[0], len(df))

        # Adjust xtick labels if the bars are negatives
        xtick_label_position = kwargs.get("xtick_label_position")
        if xtick_label_position:
            self.chart.x_axis.tickLblPos = xtick_label_position
            self.chart.x_axis.tickLblSkip = 2

            self.chart.series[0].invertIfNegative = True

        ytick_label_position = kwargs.get("ytick_label_position")
        if ytick_label_position:
            self.chart.y_axis.tickLblPos = ytick_label_position
            self.chart.y_axis.tickLblSkip = 2

            self.chart.series[0].invertIfNegative = True

        chart_type = kwargs.get("chart_type", "col")
        self.chart.type = chart_type

        super()._plot(df, ws, **kwargs)


class StackedBarChart(Chart):
    """
    A chart class for creating stacked bar charts in Excel using data from a pandas DataFrame.
    This class extends the `Chart` base class and provides functionality to plot stacked bar charts
    (either percent-stacked or regular stacked) in an Excel worksheet. It leverages the `openpyxl`
    library for chart creation and formatting.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        chart_type = kwargs.get("chart_type", "col")
        # grouping 'percentStacked' or 'stacked'
        chart_grouping = kwargs.get("chart_grouping", "percentStacked")
        chart_overlap = kwargs.get("chart_overlap", 100)
        chart_style = kwargs.get("chart_style", 10)
        self.chart = xl.chart.BarChart()
        self.chart.type = chart_type
        self.chart.grouping = chart_grouping
        self.chart.overlap = chart_overlap
        self.chart.style = chart_style

        self._add_data(df, ws, **kwargs)

        super()._plot(df, ws, **kwargs)


class PieChart(Chart):
    """
    A class for creating and customizing pie charts in Excel using openpyxl.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.PieChart()

        self._add_data(df, ws, **kwargs)

        # if any kwargs is provided (even if it is None) it takes precedence over defaults: config
        show_percentage = (
            kwargs.get("show_percentage")
            if "show_percentage" in kwargs
            else self.config.SHOW_PERCENTAGE
        )
        show_categories = (
            kwargs.get("show_category")
            if "show_category" in kwargs
            else self.config.SHOW_CATEGORIES
        )
        show_legend_key = (
            kwargs.get("show_legend_key")
            if "show_legend_key" in kwargs
            else self.config.SHOW_LEGEND_KEY
        )
        show_values = (
            kwargs.get("show_values")
            if "show_values" in kwargs
            else self.config.SHOW_VALUES
        )
        show_series_name = (
            kwargs.get("show_series_name")
            if "show_series_name" in kwargs
            else self.config.SHOW_SERIES_NAME
        )

        data_labels = xl.chart.label.DataLabelList(
            showCatName=show_categories,
            showPercent=show_percentage,
            showLegendKey=show_legend_key,
            showVal=show_values,
            showSerName=show_series_name,
        )

        # Add data labels
        self.chart.dataLabels = data_labels

        if not self.config.OPENPYXL_COLORS and not kwargs.get("openpyxl_color"):
            self.excel_helper.fill_data_point(self.chart.series[0], len(df))

        super()._plot(df, ws, **kwargs)


class BubbleChart(Chart):
    """
    A chart class for creating bubble charts in Excel using data from a pandas DataFrame.
    This class inherits from `Chart` and provides functionality to plot a bubble chart
    in an Excel worksheet. The chart uses the first three columns of the DataFrame as
    the x-axis, y-axis, and bubble size, respectively. Each row in the DataFrame is

    plotted as a separate series in the bubble chart.
    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.BubbleChart()

        for idx in range(1, len(df) + 1):
            # First column is the x-axis
            xvalues = xl.chart.Reference(
                ws,
                min_col=self.min_col,
                min_row=self.min_row + idx,
                max_row=self.max_row + idx,
            )

            # Second column is the y-axis
            yvalues = xl.chart.Reference(
                ws,
                min_col=self.min_col + 1,
                min_row=self.min_row + idx,
                max_row=self.max_row + idx,
            )

            # Third column is the bubble size
            size = xl.chart.Reference(
                ws,
                min_col=self.min_col + 2,
                min_row=self.min_row + idx,
                max_row=self.max_row + idx,
            )
            series = xl.chart.Series(
                values=yvalues,
                xvalues=xvalues,
                zvalues=size,
                title=ws.cell(row=self.min_row + idx, column=self.min_col - 1).value,
            )
            self.chart.series.append(series)

        super()._plot(df, ws, **kwargs)


class AreaChart(Chart):
    """
    A class for creating and plotting area charts in Excel worksheets.
    This class extends the `Chart` base class and provides functionality to plot area charts
    using data from a pandas DataFrame. It utilizes the `openpyxl` library to generate and
    customize area charts within Excel workbooks.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.AreaChart()

        self._add_data(df, ws, **kwargs)

        self.chart.style = kwargs.get("chart_style", 13)

        super()._plot(df, ws, **kwargs)


class ScatterChart(Chart):
    """
    ScatterChart(config=Config(), excel_helper=None)
    A chart class for creating and customizing scatter plots in Excel using openpyxl.
    This class extends the base `Chart` class and provides methods to add data from a pandas DataFrame
    to an Excel worksheet and generate a scatter chart. It supports flexible data selection, including
    adjacent and non-adjacent columns, and allows customization of chart appearance such as marker symbols,
    colors, and styles.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    See Also
    --------
    Chart : Base class for all chart types.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)

    # TODO possibly violate DRY (do not repeat yourself)
    def _add_data(self, df, ws, from_rows=False, **kwargs):
        """Add data to chart"""
        if from_rows:
            raise NotImplementedError(
                "The option `from_row` has not been implemented yet"
            )

        xvalues = xl.chart.Reference(
            ws,
            min_row=self.min_row + 1,
            max_row=self.max_row,
            min_col=1,
            max_col=1,
        )
        data_column_start = kwargs.get("data_column_start")
        data_column_end = kwargs.get("data_column_end")
        # If column start is given but not column end, set column end to column start
        data_column_end = data_column_end or data_column_start
        data_columns = kwargs.get("data_columns")

        if data_column_start is not None:
            for idx in range(data_column_start, data_column_end + 1):
                yvalues = xl.chart.Reference(
                    ws, min_row=self.min_row, max_row=self.max_row, min_col=idx
                )
                series = xl.chart.Series(yvalues, xvalues, title_from_data=True)
                self.chart.series.append(series)

            # Use pyxl indexing (starting from 1)
            self.data_columns = list(range(data_column_start, data_column_end + 1))

        elif data_columns is not None:
            # Create each series separately and add it to the figure since columns are not adjacent
            for col in data_columns:
                yvalues = xl.chart.Reference(
                    ws,
                    min_row=self.min_row,
                    max_row=self.max_row,
                    min_col=col,
                    max_col=col,
                )

                series = xl.chart.Series(yvalues, xvalues, title_from_data=True)

                self.chart.series.append(series)

            self.data_columns = list(data_columns)

            if df.iloc[:, data_columns[-1] - 1].nunique() == 1:
                self.ref_column = data_columns[-1]

            # Terminate to avoid adding data to the chart again
            return
        else:
            #
            for idx in range(self.min_col, self.max_col + 1):
                yvalues = xl.chart.Reference(
                    ws, min_row=self.min_row, max_row=self.max_row, min_col=idx
                )
                series = xl.chart.Series(yvalues, xvalues, title_from_data=True)
                self.chart.series.append(series)

            self.data_columns = list(range(self.min_col, self.max_col + 1))

            if df.iloc[:, self.max_col - 1].nunique() == 1:
                self.ref_column = self.max_col

    def plot(self, df, ws, **kwargs):
        self.write_dataframe(df, ws, **kwargs)

        self.chart = xl.chart.ScatterChart()

        self._add_data(df, ws, **kwargs)

        chart_style = kwargs.get("chart_style", 13)
        self.chart.style = chart_style

        nofill = kwargs.get("nofill")
        nofill = True if nofill is None else nofill
        kwargs["nofill"] = nofill

        super()._plot(df, ws, **kwargs)

        for idx, series in enumerate(self.chart.series):
            # Recycle colors if required
            marker_color = self.config.PRIMARY_COLORS[
                idx % len(self.config.PRIMARY_COLORS)
            ]

            marker_symbols = kwargs.get("marker_symbols")
            marker_symbol = (
                marker_symbols[idx % len(marker_symbols)]
                if marker_symbols
                else self.config.MARKER_SYMOBLS[idx % len(self.config.MARKER_SYMOBLS)]
            )

            marker_size = kwargs.get("marker_size")
            self.excel_helper.set_marker_graphical_properties(
                series,
                symbol=marker_symbol,
                size=marker_size,
                color=marker_color,
            )

            self.excel_helper.set_line_graphical_properties(series, nofill=nofill)


# NOTE `data_columns` will not work with TwoAxesChart
# TODO implement `data_columns`
class TwoAxesChart(Chart):
    """
    Chart with Two Y-Axes for Excel Worksheets.

    This class extends the `Chart` class to support plotting charts with two y-axes
    (primary and secondary) in Excel worksheets using openpyxl. It is designed to
    handle dataframes with multiple data columns, plotting the first data column on
    the primary y-axis and the remaining columns on the secondary y-axis. The class
    manages chart configuration, data assignment, and axis properties for both axes.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    Attributes
    ----------
    chart : openpyxl.chart.Chart or None
        The main chart object, representing the combined chart with two y-axes.
    chart1 : openpyxl.chart.Chart or None
        The chart object for the primary y-axis.
    chart2 : openpyxl.chart.Chart or None
        The chart object for the secondary y-axis.

    See Also
    --------
    Chart : Base class for all chart types.

    Notes
    -----
    The first data column is plotted on the primary y-axis; additional columns
    are plotted on the secondary y-axis.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)
        self.chart = None
        self.chart1 = None
        self.chart2 = None

    def plot(self, df, ws, **kwargs):

        self.write_dataframe(df, ws, **kwargs)

        # Keep chart settings that should not be set twice
        title = kwargs.get("title")
        show_legend = kwargs.get("show_legend")
        legend_position = kwargs.get("legend_position")
        chart_position = kwargs.get("chart_position")
        xlabel = kwargs.get("xlabel")
        kwargs["title"] = None
        kwargs["show_legend"] = None
        kwargs["legend_position"] = None
        # To prevent adding the chart to the sheet more than once
        kwargs["chart_position"] = "Unknown"
        kwargs["suppress_chart_position_warning"] = True
        kwargs["xlabel"] = None

        # Leave settings that need to be applied to both figure if present:
        # width, height, openpyxl_color, nofill, etc

        # Add data to chart1
        data_columns = kwargs.get("data_columns")
        # Default to second column as the first column is the category
        kwargs["data_columns"] = [data_columns[0]] if data_columns else [2]
        kwargs["set_categories"] = True

        # Set the parent self.chart to delegate to the parent class
        self.chart = self.chart1

        self._add_data(
            df,
            ws,
            **kwargs,
        )

        # Finalize the plot setting for chart1
        # super()._plot(df, ws, **kwargs)

        # Add data to chart 2
        kwargs["data_columns"] = (
            data_columns[1:]
            if data_columns
            else [i for i in range(3, len(df.columns) + 1)]
        )
        kwargs["set_categories"] = False  # Set to False to avoid duplication

        # Reset the chart of the super class to delegate to the parent class
        self.chart = self.chart2

        # Options for chart 2
        kwargs["chart_style"] = kwargs.get("chart_style2")
        kwargs["chart_shape"] = kwargs.get("chart_shape2")
        kwargs["ylabel"] = kwargs.get("ylabel2")
        kwargs["ylim"] = kwargs.get("ylim2")
        kwargs["y_orientation"] = kwargs.get("y_orientation2")
        kwargs["yaxis_major_unit"] = kwargs.get("yaxis_major_unit2")
        kwargs["y_log_base"] = kwargs.get("y_log_base2")

        self._add_data(df, ws, **kwargs)

        # Finalize the plot setting for chart2
        super()._plot(df, ws, **kwargs)

        # self.chart2.y_axis.title = kwargs.get("ylabel2")

        # Set the second y-axis
        self.chart2.y_axis.axId = 200
        self.chart2.y_axis.crosses = "max"

        # Concat the two charts
        # At this point you need to use `self.chart` of the super class
        self.chart1 += self.chart2
        self.chart = self.chart1

        # kwargs["ylabel"] = kwargs.get("ylabel") or kwargs.get("ylabel1")

        kwargs["chart_style"] = kwargs.get("chart_style1")
        kwargs["chart_shape"] = kwargs.get("chart_shape1")
        kwargs["ylabel"] = kwargs.get("ylabel1")
        kwargs["ylim"] = kwargs.get("ylim1")
        kwargs["y_orientation"] = kwargs.get("y_orientation1")
        kwargs["yaxis_major_unit"] = kwargs.get("yaxis_major_unit1")
        kwargs["y_log_base"] = kwargs.get("y_log_base1")
        # Restore general settings
        kwargs["title"] = title
        kwargs["show_legend"] = show_legend
        kwargs["legend_position"] = legend_position
        kwargs["chart_position"] = chart_position or self.config.CHART_POSITION
        kwargs["suppress_chart_position_warning"] = False
        kwargs["xlabel"] = xlabel

        # Delegate
        super()._plot(df=df, ws=ws, **kwargs)


class BarLineChart(TwoAxesChart):
    """
    A chart class for creating combined Bar and Line charts in Excel using openpyxl.
    This class extends `TwoAxesChart` and provides functionality to plot a bar chart
    (as the primary axis) and a line chart (as the secondary axis) on the same Excel worksheet.
    It allows customization of chart styles, shapes, line properties, and marker properties.

    Parameters
    ----------
    config : Config, optional
            Configuration object containing chart and general Excel settings.
    excel_helper : ExcelHelper, optional
            Helper object for Excel-specific operations.

    Attributes
    ----------
    chart : None or openpyxl.chart.Chart
        Placeholder for the combined chart object.
    chart1 : openpyxl.chart.BarChart
        The bar chart instance (primary axis).
    chart2 : openpyxl.chart.LineChart
        The line chart instance (secondary axis).

    See Also
    --------
    Chart : Base class for all chart types.
    TwoAxesChart: Parent class.
    """

    def __init__(self, config=Config(), excel_helper=None):
        super().__init__(config, excel_helper)
        self.chart = None
        self.chart1 = xl.chart.BarChart()
        self.chart2 = xl.chart.LineChart()

    def plot(self, df, ws, **kwargs):

        # Override default
        openpyxl_color = kwargs.get("openpyxl_color")
        kwargs["openpyxl_color"] = kwargs.get("openpyxl_color", True)

        self.chart1.type = "col"
        self.chart1.style = kwargs.get("bar_chart_style", 10)
        self.chart1.shape = kwargs.get("chart_shape", 4)

        #
        super().plot(df, ws, **kwargs)
        self.chart1.y_axis.majorGridlines = None

        self.chart2.style = kwargs.get("line_chart_style", 10)
        line_width = kwargs.get("line_width", 1.5)
        line_style = kwargs.get("line_style", "sysDash")
        smooth = kwargs.get("smooth", True)
        kwargs["openpyxl_color"] = openpyxl_color

        for idx, series in enumerate(self.chart2.series):
            idx += 1
            color = (
                None  # Use default openpyxl color
                if self.config.OPENPYXL_COLORS or kwargs.get("openpyxl_color")
                # Recycle colors if needed
                else self.config.PRIMARY_COLORS[(idx) % len(self.config.PRIMARY_COLORS)]
            )

            # Use line color if it is user-defined
            if not kwargs.get("line_color") and idx == 0:
                color = "107C10"
            else:
                color = kwargs.get("line_color") or color

            self.excel_helper.set_line_graphical_properties(
                series,
                line_width,
                style=line_style,
                color=color,
                smooth=smooth,
            )

            marker_symbol = kwargs.get("marker_symbol", "circle")
            marker_size = kwargs.get("marker_size", 6)
            self.excel_helper.set_marker_graphical_properties(
                series, marker_symbol, marker_size, color
            )


# TODO Coming soon ...
class LineLineChart(TwoAxesChart):

    def __init__(self, config=Config(), excel_helper=None):

        raise NotImplementedError("Not implemented yet")
